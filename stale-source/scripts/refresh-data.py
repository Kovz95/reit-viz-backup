#!/usr/bin/env python3
"""
REIT Data Refresh Script
========================
Extracts ticker data from the Excel (.xlsb) workbook and writes JSON files
for the REIT Viz dashboard.

Supports incremental mode: only reprocesses sheets whose data has changed
since the last run (based on row-level fingerprints stored in a manifest).

Usage:
    python3 scripts/refresh-data.py [path_to_xlsb]              # incremental (default)
    python3 scripts/refresh-data.py [path_to_xlsb] --full       # force full reprocess
    python3 scripts/refresh-data.py [path_to_xlsb] --incremental  # explicit incremental

Output:
    data/dates.json            - Array of date strings
    data/tickers.json          - Array of ticker metadata objects
    data/events.json           - Ex-dividend and earnings dates per ticker
    data/tickers/XYZ.json      - Per-ticker metric data (run-length encoded)
    data/sources.json          - Workbook source tracking per ticker
    data/.ingest-manifest.json - Sheet fingerprints for incremental mode
"""

import sys
import os
import json
import time
import hashlib
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed

# Import the fast XLSB parser (same directory)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from fast_xlsb import FastXlsbReader

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

# ---------- CLI ----------

args = sys.argv[1:]
flags = [a for a in args if a.startswith("--")]
positional = [a for a in args if not a.startswith("--")]

FORCE_FULL = "--full" in flags
# Default to incremental (no flag needed)

EXCEL_PATH = positional[0] if positional else os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "..", "REITS-total-universe-Copy.xlsb"
)

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
MANIFEST_PATH = os.path.join(DATA_DIR, ".ingest-manifest.json")

# ---------- Row mapping ----------

ROW_MAP = {
    9:  "close",
    10: "open",
    11: "low",
    12: "high",
    21: "EPS FY1",
    22: "EPS FY2",
    23: "EBITDA FY1",
    24: "EBITDA FY2",
    25: "FFO FY1",
    26: "FFO FY2",
    27: "AFFO FY1",
    28: "AFFO FY2",
    29: "Sales FY1",
    30: "Sales FY2",
    37: "EPS LTM",
    38: "Sales LTM",
    39: "EBITDA LTM",
    40: "FFO LTM",
    41: "AFFO LTM",
    42: "EPS FY0",
    47: "FFO FY0",
    48: "AFFO FY0",
    49: "Dividend",
    50: "Enterprise Value",
    51: "52wk High",
    52: "52wk Low",
    66: "1Y Price Chg%",
    67: "6M Price Chg%",
    68: "3M Price Chg%",
    69: "1M Price Chg%",
    70: "Short Interest%",
    71: "Buy Ratings",
    72: "Hold Ratings",
    74: "Sell Ratings",
    76: "Bull%",
    77: "Bear%",
    78: "FY1 EPS Growth",
    79: "FY2 EPS Growth",
    86: "FY1 FFO Growth",
    87: "FY2 FFO Growth",
    88: "FY1 AFFO Growth",
    89: "FY2 AFFO Growth",
    92: "% off 52wk High",
    93: "% off 52wk Low",
    95: "P/E LTM",
    96: "P/E FY2",
    97: "P/S LTM",
    98: "P/S FY2",
    101: "EV/EBITDA LTM",
    102: "EV/EBITDA FY2",
    109: "P/FFO LTM",
    110: "P/FFO FY2",
    111: "P/AFFO LTM",
    112: "P/AFFO FY2",
    113: "FFO Yield LTM",
    114: "FFO Yield FY2",
    115: "AFFO Yield LTM",
    116: "AFFO Yield FY2",
    127: "Dividend Yield",
}

NEEDED_ROWS = {8} | set(ROW_MAP.keys())
MAX_ROW = max(NEEDED_ROWS)

# Rows used for fingerprinting (dates + close + a few estimates)
# Reading just these rows is enough to detect any meaningful data change.
FINGERPRINT_ROWS = {8, 9, 25, 109, 127}


# ---------- Helpers ----------

def parse_date(val):
    """Convert date string from Excel to YYYY-MM-DD format."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def run_length_encode(values):
    """Encode null runs as ~N for compact storage."""
    encoded = []
    null_count = 0
    for v in values:
        if v is None:
            null_count += 1
        else:
            if null_count > 0:
                encoded.append(f"~{null_count}")
                null_count = 0
            if isinstance(v, float):
                encoded.append(round(v, 4))
            else:
                encoded.append(v)
    if null_count > 0:
        encoded.append(f"~{null_count}")
    return encoded


def compute_sheet_fingerprint(row_values_dict):
    """Compute a SHA-256 fingerprint from a dict of row_num -> [values]."""
    h = hashlib.sha256()
    for row_num in sorted(row_values_dict.keys()):
        vals = row_values_dict[row_num]
        # Hash the last 20 values (most recent data — where changes happen)
        tail = vals[-20:] if len(vals) > 20 else vals
        h.update(f"{row_num}:{tail}".encode())
    return h.hexdigest()[:16]  # 16 hex chars is enough


def load_manifest():
    """Load the previous ingest manifest."""
    if os.path.exists(MANIFEST_PATH):
        try:
            return json.load(open(MANIFEST_PATH))
        except Exception:
            pass
    return {"workbook": None, "workbook_mtime": None, "sheets": {}}


def save_manifest(manifest):
    """Save the ingest manifest."""
    os.makedirs(os.path.dirname(MANIFEST_PATH), exist_ok=True)
    with open(MANIFEST_PATH, 'w') as f:
        json.dump(manifest, f, separators=(',', ':'), indent=2)


# ---------- Fingerprinting ----------

def fingerprint_single_sheet(args):
    """Read only fingerprint rows from a single sheet (subprocess). Returns (sheet_name, ticker, fingerprint)."""
    excel_path, sheet_name = args
    ticker = sheet_name.replace("-US_mktdata", "").replace("_mktdata", "").strip()
    ext = os.path.splitext(excel_path)[1].lower()

    # Import fast_xlsb in subprocess
    import sys as _sys
    _sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from fast_xlsb import FastXlsbReader as _FastReader

    try:
        row_values = {}
        if ext in ('.xlsx', '.xls', '.xlsm'):
            wb_ox = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb_ox[sheet_name]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                row_num = i + 1
                if row_num in FINGERPRINT_ROWS:
                    row_values[row_num] = [str(v) for v in row[1:] if v is not None][-20:]
                if row_num > max(FINGERPRINT_ROWS):
                    break
            wb_ox.close()
        else:
            # Use fast binary parser — read only fingerprint rows
            # FINGERPRINT_ROWS uses 1-indexed rows; FastXlsbReader uses 0-indexed
            needed_0 = {r - 1 for r in FINGERPRINT_ROWS}
            with _FastReader(excel_path) as reader:
                rows = reader.read_sheet_rows(sheet_name, needed_rows=needed_0, max_row=max(needed_0))
                for r0, vals in rows.items():
                    r1 = r0 + 1
                    if r1 in FINGERPRINT_ROWS:
                        row_values[r1] = [str(v) for v in vals[1:] if v is not None][-20:]

        fp = compute_sheet_fingerprint(row_values)
        return (sheet_name, ticker, fp)
    except Exception as e:
        return (sheet_name, ticker, f"ERROR:{e}")


# ---------- Full sheet processing ----------

def process_single_sheet(args):
    """Process a single mktdata sheet in a subprocess.
    Returns (ticker, dates, num_dates, available_metrics, file_size, fingerprint).
    On error returns 7-tuple with error string at the end."""
    excel_path, sheet_name, tickers_dir = args
    ticker = sheet_name.replace("-US_mktdata", "").replace("_mktdata", "").strip()

    ext = os.path.splitext(excel_path)[1].lower()
    use_openpyxl = ext in ('.xlsx', '.xls', '.xlsm')

    # Import fast_xlsb in subprocess
    import sys as _sys
    _sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from fast_xlsb import FastXlsbReader as _FastReader

    try:
        fp_rows = {}  # for fingerprinting
        if use_openpyxl:
            wb_ox = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb_ox[sheet_name]
            rows_data = {}
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                row_num = i + 1
                if row_num in FINGERPRINT_ROWS:
                    fp_rows[row_num] = [str(v) for v in row[1:] if v is not None][-20:]
                if row_num == 8:
                    date_vals = list(row[1:])
                    dates = [parse_date(v) for v in date_vals]
                    while dates and dates[-1] is None:
                        dates.pop()
                    rows_data['_dates'] = dates
                elif row_num in ROW_MAP:
                    metric_name = ROW_MAP[row_num]
                    vals = list(row[1:])
                    clean_vals = []
                    for v in vals:
                        if v is None or v == '' or v == ' ' or v is False:
                            clean_vals.append(None)
                        else:
                            try:
                                clean_vals.append(float(v))
                            except (ValueError, TypeError):
                                clean_vals.append(None)
                    rows_data[metric_name] = clean_vals
                if row_num > MAX_ROW:
                    break
            wb_ox.close()
        else:
            # Use fast binary parser — read only the rows we need
            needed_0 = {r - 1 for r in NEEDED_ROWS}  # 0-indexed
            fp_needed_0 = {r - 1 for r in FINGERPRINT_ROWS}
            all_needed_0 = needed_0 | fp_needed_0
            with _FastReader(excel_path) as reader:
                all_rows = reader.read_sheet_rows(sheet_name, needed_rows=all_needed_0, max_row=MAX_ROW - 1)

            rows_data = {}
            # Build fingerprint rows (1-indexed)
            for r0, vals in all_rows.items():
                r1 = r0 + 1
                if r1 in FINGERPRINT_ROWS:
                    fp_rows[r1] = [str(v) for v in vals[1:] if v is not None][-20:]

            # Row 8 (0-indexed: 7) = dates
            if 7 in all_rows:
                date_vals = all_rows[7][1:]  # skip col A
                dates = [parse_date(v) for v in date_vals]
                while dates and dates[-1] is None:
                    dates.pop()
                rows_data['_dates'] = dates

            # Metric rows
            for row_num_1, metric_name in ROW_MAP.items():
                r0 = row_num_1 - 1
                if r0 not in all_rows:
                    continue
                vals = all_rows[r0][1:]  # skip col A
                clean_vals = []
                for v in vals:
                    if v is None or v == '' or v == ' ' or v is False:
                        clean_vals.append(None)
                    else:
                        try:
                            clean_vals.append(float(v))
                        except (ValueError, TypeError):
                            clean_vals.append(None)
                rows_data[metric_name] = clean_vals

        fp = compute_sheet_fingerprint(fp_rows)

        if '_dates' not in rows_data:
            return (ticker, None, 0, [], 0, fp)

        dates = rows_data['_dates']
        num_dates = len(dates)

        ticker_data = {}
        available_metrics = []
        for metric_name, values in rows_data.items():
            if metric_name == '_dates':
                continue
            trimmed = values[:num_dates]
            while len(trimmed) < num_dates:
                trimmed.append(None)
            has_data = any(v is not None for v in trimmed)
            if has_data:
                ticker_data[metric_name] = run_length_encode(trimmed)
                available_metrics.append(metric_name)

        # Save ticker JSON
        ticker_file = os.path.join(tickers_dir, f"{ticker}.json")
        with open(ticker_file, 'w') as f:
            json.dump(ticker_data, f, separators=(',', ':'))

        file_size = os.path.getsize(ticker_file)
        return (ticker, dates, num_dates, available_metrics, file_size, fp)

    except Exception as e:
        return (ticker, None, 0, [], 0, "", str(e))


# ---------- Main ----------

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    t_start = time.time()
    wb_mtime = os.path.getmtime(EXCEL_PATH)
    wb_size = os.path.getsize(EXCEL_PATH)
    mode_label = "FULL" if FORCE_FULL else "INCREMENTAL"
    print(f"[{mode_label}] Reading workbook: {EXCEL_PATH}")
    print(f"  File size: {wb_size / (1024*1024):.1f} MB, mtime: {datetime.fromtimestamp(wb_mtime).isoformat()}")

    ext = os.path.splitext(EXCEL_PATH)[1].lower()
    use_openpyxl = ext in ('.xlsx', '.xls', '.xlsm')

    if use_openpyxl:
        print(f"  Using openpyxl for {ext} format")
        wb_ox = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        sheets = wb_ox.sheetnames
        class SheetAdapter:
            def __init__(self, ws):
                self._ws = ws
            def __enter__(self):
                return self
            def __exit__(self, *a):
                pass
            def rows(self):
                for row in self._ws.iter_rows(values_only=False):
                    class Cell:
                        def __init__(self, v):
                            self.v = v
                    yield [Cell(c.value) for c in row]
        class WbAdapter:
            def __init__(self, ox):
                self._ox = ox
                self.sheets = ox.sheetnames
            def get_sheet(self, name):
                return SheetAdapter(self._ox[name])
        wb = WbAdapter(wb_ox)
    else:
        print(f"  Using fast_xlsb binary parser for {ext} format")
        fast_reader = FastXlsbReader(EXCEL_PATH)
        sheets = fast_reader.sheet_names

    # Find mktdata sheets
    mktdata_sheets = [s for s in sheets if '_mktdata' in s.lower()]
    print(f"Found {len(mktdata_sheets)} ticker sheets")

    # ---------- Extract Tickerlist ----------
    t_tl = time.time()
    print("Extracting tickerlist...")
    ticker_meta = {}
    if use_openpyxl:
        with wb.get_sheet('Tickerlist') as sheet:
            for i, row in enumerate(sheet.rows()):
                if i == 0:
                    continue  # header
                vals = [c.v for c in row[:8]]
                if not vals[0]:
                    continue
                raw_ticker = str(vals[0]).replace("-US", "").strip()
                ticker_meta[raw_ticker] = {
                    "ticker": raw_ticker,
                    "name": str(vals[1] or "").strip(),
                    "economy": str(vals[2] or "").strip(),
                    "sector": str(vals[3] or "").strip(),
                    "subsector": str(vals[4] or "").strip(),
                    "industryGroup": str(vals[5] or "").strip(),
                    "industry": str(vals[6] or "").strip(),
                    "subindustry": str(vals[7] or "Other").strip(),
                }
    else:
        # Use fast_xlsb to read Tickerlist — read enough rows
        tl_rows = fast_reader.read_sheet_rows('Tickerlist', needed_rows=None, max_row=200)
        for r0 in sorted(tl_rows.keys()):
            if r0 == 0:
                continue  # header
            vals = tl_rows[r0]
            if len(vals) < 8 or not vals[0]:
                continue
            raw_ticker = str(vals[0]).replace("-US", "").strip()
            ticker_meta[raw_ticker] = {
                "ticker": raw_ticker,
                "name": str(vals[1] or "").strip() if len(vals) > 1 else "",
                "economy": str(vals[2] or "").strip() if len(vals) > 2 else "",
                "sector": str(vals[3] or "").strip() if len(vals) > 3 else "",
                "subsector": str(vals[4] or "").strip() if len(vals) > 4 else "",
                "industryGroup": str(vals[5] or "").strip() if len(vals) > 5 else "",
                "industry": str(vals[6] or "").strip() if len(vals) > 6 else "",
                "subindustry": str(vals[7] or "Other").strip() if len(vals) > 7 else "Other",
            }
    print(f"  {len(ticker_meta)} tickers in tickerlist ({time.time()-t_tl:.1f}s)")

    # ---------- Extract Events ----------
    t_ev = time.time()
    print("Extracting events...")
    events = {}

    for sheet_name, event_key in [("Ex dividend dates", "ex_dividend"), ("Earnings report dates", "earnings")]:
        if sheet_name not in sheets:
            print(f"  WARNING: Sheet '{sheet_name}' not found, skipping")
            continue
        if use_openpyxl:
            with wb.get_sheet(sheet_name) as sheet:
                for i, row in enumerate(sheet.rows()):
                    if i == 0:
                        continue
                    vals = [c.v for c in row]
                    if not vals[0]:
                        continue
                    ticker = str(vals[0]).replace("-US^", "").replace("-US", "").strip()
                    dates_list = []
                    for v in vals[1:]:
                        d = parse_date(v)
                        if d:
                            dates_list.append(d)
                    if ticker not in events:
                        events[ticker] = {}
                    events[ticker][event_key] = dates_list
        else:
            # Use fast_xlsb — events can have many rows, read all up to 200
            ev_rows = fast_reader.read_sheet_rows(sheet_name, needed_rows=None, max_row=200)
            for r0 in sorted(ev_rows.keys()):
                if r0 == 0:
                    continue
                vals = ev_rows[r0]
                if not vals or not vals[0]:
                    continue
                ticker = str(vals[0]).replace("-US^", "").replace("-US", "").strip()
                dates_list = []
                for v in vals[1:]:
                    d = parse_date(v)
                    if d:
                        dates_list.append(d)
                if ticker not in events:
                    events[ticker] = {}
                events[ticker][event_key] = dates_list

    print(f"  Events extracted ({time.time()-t_ev:.1f}s)")

    # Close the workbook handle before parallel processing
    if use_openpyxl:
        wb_ox.close()
    else:
        fast_reader.close()

    # ---------- Incremental: Fingerprint Phase ----------
    import multiprocessing
    num_workers = min(multiprocessing.cpu_count(), 4)
    tickers_dir = os.path.join(DATA_DIR, "tickers")
    os.makedirs(tickers_dir, exist_ok=True)

    manifest = load_manifest()
    old_fingerprints = manifest.get("sheets", {})

    # Load existing tickers.json for reuse of unchanged ticker metadata
    existing_tickers_map = {}
    existing_tickers_file = os.path.join(DATA_DIR, "tickers.json")
    if os.path.exists(existing_tickers_file):
        try:
            for t in json.load(open(existing_tickers_file)):
                existing_tickers_map[t["ticker"]] = t
        except Exception:
            pass

    # ── Tier 1: File-level check ──
    # If workbook mtime+size haven't changed and all ticker JSONs exist, skip everything
    skip_all = False
    if not FORCE_FULL:
        old_mtime = manifest.get("workbook_mtime")
        old_size = manifest.get("workbook_size")
        if old_mtime == wb_mtime and old_size == wb_size and len(old_fingerprints) > 0:
            # Verify all ticker JSONs still exist
            all_exist = all(
                os.path.exists(os.path.join(tickers_dir, f"{v['ticker']}.json"))
                for v in old_fingerprints.values()
            )
            if all_exist:
                skip_all = True

    if skip_all:
        print(f"\n  Workbook unchanged (same mtime + size) — skipping all {len(mktdata_sheets)} sheets")
        sheets_to_process = []
        skipped_sheets = mktdata_sheets
        new_fingerprints = {sn: old_fingerprints[sn]["fingerprint"] for sn in mktdata_sheets if sn in old_fingerprints}
    elif FORCE_FULL:
        # Skip fingerprinting — process everything
        sheets_to_process = mktdata_sheets
        new_fingerprints = {}
        skipped_sheets = []
        print(f"\n  [FULL MODE] Will process all {len(mktdata_sheets)} sheets")
    else:
        # ── Tier 2: Sheet-level fingerprinting ──
        t_fp = time.time()
        print(f"\nFingerprinting {len(mktdata_sheets)} sheets ({num_workers} workers)...")
        fp_args = [(EXCEL_PATH, sn) for sn in mktdata_sheets]

        new_fingerprints = {}
        sheets_to_process = []
        skipped_sheets = []

        with ProcessPoolExecutor(max_workers=num_workers) as pool:
            futures = {pool.submit(fingerprint_single_sheet, a): a[1] for a in fp_args}
            for future in as_completed(futures):
                sheet_name, ticker, fp = future.result()
                if fp.startswith("ERROR:"):
                    print(f"  {ticker}: fingerprint error — will reprocess")
                    sheets_to_process.append(sheet_name)
                    continue

                new_fingerprints[sheet_name] = fp
                old_fp = old_fingerprints.get(sheet_name, {}).get("fingerprint")
                existing_json = os.path.join(tickers_dir, f"{ticker}.json")

                if old_fp == fp and os.path.exists(existing_json):
                    skipped_sheets.append(sheet_name)
                else:
                    sheets_to_process.append(sheet_name)

        fp_elapsed = time.time() - t_fp
        print(f"  Fingerprinting done in {fp_elapsed:.1f}s")
        print(f"  Changed: {len(sheets_to_process)}, Unchanged: {len(skipped_sheets)}")

    # ---------- Phase 2: Process changed sheets ----------
    t_tk = time.time()
    all_dates = None
    tickers_list = []
    completed = 0

    # Map sheet_name -> fingerprint computed during processing
    processed_fingerprints = {}

    if sheets_to_process:
        print(f"\nProcessing {len(sheets_to_process)} changed sheets ({num_workers} workers)...")
        work_args = [(EXCEL_PATH, sn, tickers_dir) for sn in sheets_to_process]

        with ProcessPoolExecutor(max_workers=num_workers) as pool:
            futures = {pool.submit(process_single_sheet, a): a[1] for a in work_args}
            for future in as_completed(futures):
                completed += 1
                result = future.result()
                if len(result) > 6:
                    # Error: (ticker, None, 0, [], 0, "", err_str)
                    ticker = result[0]
                    err = result[-1]
                    print(f"  [{completed}/{len(sheets_to_process)}] {ticker} ERROR: {err}")
                    continue

                ticker, dates, num_dates, available_metrics, file_size, fp = result
                sheet_name = [sn for sn in sheets_to_process if sn.replace("-US_mktdata", "").replace("_mktdata", "").strip() == ticker][0]
                processed_fingerprints[sheet_name] = fp

                if dates is None:
                    print(f"  [{completed}/{len(sheets_to_process)}] {ticker} SKIP (no dates)")
                    continue

                print(f"  [{completed}/{len(sheets_to_process)}] {ticker}: {len(available_metrics)} metrics, {file_size//1024}KB")

                if all_dates is None or len(dates) > len(all_dates):
                    all_dates = dates

                meta = ticker_meta.get(ticker, {
                    "ticker": ticker, "name": ticker, "economy": "", "sector": "",
                    "subsector": "", "industryGroup": "", "industry": "Other", "subindustry": "Other",
                })
                meta["dates"] = num_dates
                meta["metrics"] = available_metrics
                tickers_list.append(meta)

        print(f"  Processing done ({time.time()-t_tk:.1f}s)")
    else:
        print("\n  No changed sheets — all tickers up to date!")

    # ---------- Collect unchanged tickers from existing data ----------
    processed_tickers = {t["ticker"] for t in tickers_list}

    for sheet_name in skipped_sheets:
        ticker = sheet_name.replace("-US_mktdata", "").replace("_mktdata", "").strip()
        if ticker in processed_tickers:
            continue  # Already processed (shouldn't happen)

        # Reuse existing metadata
        if ticker in existing_tickers_map:
            tickers_list.append(existing_tickers_map[ticker])
        else:
            # Ticker JSON exists but no metadata — read the file to reconstruct
            ticker_file = os.path.join(tickers_dir, f"{ticker}.json")
            if os.path.exists(ticker_file):
                try:
                    td = json.load(open(ticker_file))
                    meta = ticker_meta.get(ticker, {
                        "ticker": ticker, "name": ticker, "economy": "", "sector": "",
                        "subsector": "", "industryGroup": "", "industry": "Other", "subindustry": "Other",
                    })
                    meta["metrics"] = list(td.keys())
                    # Estimate dates from close array length (tuple format)
                    if "close" in td:
                        close_data = td["close"]
                        # Count actual data points from RLE
                        count = 0
                        for v in close_data:
                            if isinstance(v, str) and v.startswith("~"):
                                count += int(v[1:])
                            else:
                                count += 1
                        meta["dates"] = count
                    tickers_list.append(meta)
                except Exception:
                    pass

    # Update canonical dates from existing if we didn't process any sheets
    if all_dates is None:
        existing_dates_file = os.path.join(DATA_DIR, "dates.json")
        if os.path.exists(existing_dates_file):
            try:
                all_dates = json.load(open(existing_dates_file))
            except Exception:
                pass

    if all_dates is None:
        print("ERROR: No date data available")
        sys.exit(1)

    # ---------- Write Output Files ----------
    print("\nWriting output files...")

    # dates.json
    dates_file = os.path.join(DATA_DIR, "dates.json")
    with open(dates_file, 'w') as f:
        json.dump(all_dates, f, separators=(',', ':'))
    print(f"  dates.json: {len(all_dates)} dates")

    # tickers.json
    tickers_list.sort(key=lambda t: t["ticker"])
    tickers_file = os.path.join(DATA_DIR, "tickers.json")
    with open(tickers_file, 'w') as f:
        json.dump(tickers_list, f, separators=(',', ':'))
    print(f"  tickers.json: {len(tickers_list)} tickers")

    # events.json
    events_file = os.path.join(DATA_DIR, "events.json")
    with open(events_file, 'w') as f:
        json.dump(events, f, separators=(',', ':'))
    print(f"  events.json: {len(events)} tickers with events")

    # sources.json
    wb_filename = os.path.basename(EXCEL_PATH)
    upload_ts = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

    # Load existing sources to preserve upload timestamps for unchanged tickers
    sources_file = os.path.join(DATA_DIR, "sources.json")
    existing_sources = {}
    if os.path.exists(sources_file):
        try:
            existing_sources = json.load(open(sources_file))
        except Exception:
            pass

    sources = {}
    for t in tickers_list:
        ticker_sym = t["ticker"]
        ticker_file = os.path.join(DATA_DIR, "tickers", f"{ticker_sym}.json")
        file_size = os.path.getsize(ticker_file) if os.path.exists(ticker_file) else 0

        if ticker_sym in processed_tickers:
            # Freshly processed — update timestamp
            sources[ticker_sym] = {
                "workbook": wb_filename,
                "uploadedAt": upload_ts,
                "dates": t.get("dates", 0),
                "metrics": len(t.get("metrics", [])),
                "fileSize": file_size,
            }
        elif ticker_sym in existing_sources:
            # Unchanged — keep existing source info
            sources[ticker_sym] = existing_sources[ticker_sym]
            # Update file size in case it changed
            sources[ticker_sym]["fileSize"] = file_size
        else:
            sources[ticker_sym] = {
                "workbook": wb_filename,
                "uploadedAt": upload_ts,
                "dates": t.get("dates", 0),
                "metrics": len(t.get("metrics", [])),
                "fileSize": file_size,
            }

    with open(sources_file, 'w') as f:
        json.dump(sources, f, separators=(',', ':'), indent=2)
    print(f"  sources.json: {len(sources)} ticker sources tracked")

    # ---------- Update Manifest ----------
    new_manifest = {
        "workbook": os.path.basename(EXCEL_PATH),
        "workbook_mtime": wb_mtime,
        "workbook_size": wb_size,
        "last_run": datetime.now().isoformat(),
        "mode": mode_label,
        "sheets": {},
    }

    # For processed sheets, use the fingerprint computed during processing
    for sn in sheets_to_process:
        ticker = sn.replace("-US_mktdata", "").replace("_mktdata", "").strip()
        # Prefer fingerprint from processing (always available now), fall back to
        # incremental-phase fingerprint
        fp = processed_fingerprints.get(sn) or new_fingerprints.get(sn) or ""
        new_manifest["sheets"][sn] = {
            "fingerprint": fp,
            "ticker": ticker,
            "processed_at": datetime.now().isoformat(),
        }

    # For skipped sheets, carry forward their fingerprints
    for sn in skipped_sheets:
        ticker = sn.replace("-US_mktdata", "").replace("_mktdata", "").strip()
        fp = new_fingerprints.get(sn) or old_fingerprints.get(sn, {}).get("fingerprint", "")
        new_manifest["sheets"][sn] = {
            "fingerprint": fp,
            "ticker": ticker,
            "processed_at": old_fingerprints.get(sn, {}).get("processed_at", datetime.now().isoformat()),
        }

    save_manifest(new_manifest)
    print(f"  manifest: {len(new_manifest['sheets'])} sheet fingerprints saved")

    # ---------- Summary ----------
    total_size = 0
    for root, dirs, files in os.walk(DATA_DIR):
        for file in files:
            total_size += os.path.getsize(os.path.join(root, file))

    elapsed = time.time() - t_start
    print(f"\nTotal data size: {total_size / (1024*1024):.1f} MB")
    print(f"Processed: {len(sheets_to_process)}, Skipped: {len(skipped_sheets)}, Total: {len(tickers_list)}")
    print(f"Done in {elapsed:.1f}s!")


if __name__ == "__main__":
    main()
