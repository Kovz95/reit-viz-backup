#!/usr/bin/env python3
"""
Multi-Workbook Merge Script
============================
Two-phase upload: preview (detect conflicts) then merge (apply with resolutions).

Usage:
    # Phase 1: Preview — parse new workbook into temp dir, detect conflicts
    python3 scripts/merge-workbook.py preview <path_to_xlsb> <temp_dir>
    → Outputs JSON to stdout: { newTickers: [...], conflicts: [...], newEvents: int }

    # Phase 2: Merge — apply temp data into main data/ dir with conflict resolutions
    python3 scripts/merge-workbook.py merge <temp_dir> <resolutions_json>
    → resolutions_json: JSON string like {"SPG":"overwrite","O":"keep"}
    → Merges temp data into data/, respecting resolutions

    # Phase 3: Cleanup temp dir
    Caller removes temp_dir after merge.
"""

import sys
import os
import json
import time
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed

# Import the fast XLSB parser (same directory)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from fast_xlsb import FastXlsbReader

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

# ── Config (same as refresh-data.py) ──

ROW_MAP = {
    9:  "close", 10: "open", 11: "low", 12: "high",
    21: "EPS FY1", 22: "EPS FY2", 23: "EBITDA FY1", 24: "EBITDA FY2",
    25: "FFO FY1", 26: "FFO FY2", 27: "AFFO FY1", 28: "AFFO FY2",
    29: "Sales FY1", 30: "Sales FY2",
    37: "EPS LTM", 38: "Sales LTM", 39: "EBITDA LTM", 40: "FFO LTM", 41: "AFFO LTM",
    42: "EPS FY0", 47: "FFO FY0", 48: "AFFO FY0",
    49: "Dividend", 50: "Enterprise Value", 51: "52wk High", 52: "52wk Low",
    66: "1Y Price Chg%", 67: "6M Price Chg%", 68: "3M Price Chg%", 69: "1M Price Chg%",
    70: "Short Interest%", 71: "Buy Ratings", 72: "Hold Ratings", 74: "Sell Ratings",
    76: "Bull%", 77: "Bear%",
    78: "FY1 EPS Growth", 79: "FY2 EPS Growth",
    86: "FY1 FFO Growth", 87: "FY2 FFO Growth", 88: "FY1 AFFO Growth", 89: "FY2 AFFO Growth",
    92: "% off 52wk High", 93: "% off 52wk Low",
    95: "P/E LTM", 96: "P/E FY2", 97: "P/S LTM", 98: "P/S FY2",
    101: "EV/EBITDA LTM", 102: "EV/EBITDA FY2",
    109: "P/FFO LTM", 110: "P/FFO FY2", 111: "P/AFFO LTM", 112: "P/AFFO FY2",
    113: "FFO Yield LTM", 114: "FFO Yield FY2", 115: "AFFO Yield LTM", 116: "AFFO Yield FY2",
    127: "Dividend Yield",
}

NEEDED_ROWS = {8} | set(ROW_MAP.keys())
MAX_ROW = max(NEEDED_ROWS)

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")


def parse_date(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def run_length_encode(values):
    encoded = []
    null_count = 0
    for v in values:
        if v is None:
            null_count += 1
        else:
            if null_count > 0:
                encoded.append(f"~{null_count}")
                null_count = 0
            if isinstance(v, float):
                encoded.append(round(v, 4))
            else:
                encoded.append(v)
    if null_count > 0:
        encoded.append(f"~{null_count}")
    return encoded


def process_single_sheet(args):
    """Process a single mktdata sheet in a subprocess."""
    excel_path, sheet_name, tickers_dir = args
    ticker = sheet_name.replace("-US_mktdata", "").replace("_mktdata", "").strip()

    ext = os.path.splitext(excel_path)[1].lower()
    use_openpyxl = ext in ('.xlsx', '.xls', '.xlsm')

    try:
        if use_openpyxl:
            wb_ox = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb_ox[sheet_name]
            rows_data = {}
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                row_num = i + 1
                if row_num == 8:
                    date_vals = list(row[1:])
                    dates = [parse_date(v) for v in date_vals]
                    while dates and dates[-1] is None:
                        dates.pop()
                    rows_data['_dates'] = dates
                elif row_num in ROW_MAP:
                    metric_name = ROW_MAP[row_num]
                    vals = list(row[1:])
                    clean_vals = []
                    for v in vals:
                        if v is None or v == '' or v == ' ' or v is False:
                            clean_vals.append(None)
                        else:
                            try:
                                clean_vals.append(float(v))
                            except (ValueError, TypeError):
                                clean_vals.append(None)
                    rows_data[metric_name] = clean_vals
                if row_num > MAX_ROW:
                    break
            wb_ox.close()
        else:
            # Import fast_xlsb in subprocess
            import sys as _sys
            _sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
            from fast_xlsb import FastXlsbReader as _FastReader

            needed_0 = {r - 1 for r in NEEDED_ROWS}  # 0-indexed
            with _FastReader(excel_path) as reader:
                all_rows = reader.read_sheet_rows(sheet_name, needed_rows=needed_0, max_row=MAX_ROW - 1)

            rows_data = {}
            # Row 8 (0-indexed: 7) = dates
            if 7 in all_rows:
                date_vals = all_rows[7][1:]  # skip col A
                dates = [parse_date(v) for v in date_vals]
                while dates and dates[-1] is None:
                    dates.pop()
                rows_data['_dates'] = dates

            # Metric rows
            for row_num_1, metric_name in ROW_MAP.items():
                r0 = row_num_1 - 1
                if r0 not in all_rows:
                    continue
                vals = all_rows[r0][1:]  # skip col A
                clean_vals = []
                for v in vals:
                    if v is None or v == '' or v == ' ' or v is False:
                        clean_vals.append(None)
                    else:
                        try:
                            clean_vals.append(float(v))
                        except (ValueError, TypeError):
                            clean_vals.append(None)
                rows_data[metric_name] = clean_vals

        if '_dates' not in rows_data:
            return (ticker, None, 0, [], 0)

        dates = rows_data['_dates']
        num_dates = len(dates)

        ticker_data = {}
        available_metrics = []
        for metric_name, values in rows_data.items():
            if metric_name == '_dates':
                continue
            trimmed = values[:num_dates]
            while len(trimmed) < num_dates:
                trimmed.append(None)
            has_data = any(v is not None for v in trimmed)
            if has_data:
                ticker_data[metric_name] = run_length_encode(trimmed)
                available_metrics.append(metric_name)

        ticker_file = os.path.join(tickers_dir, f"{ticker}.json")
        with open(ticker_file, 'w') as f:
            json.dump(ticker_data, f, separators=(',', ':'))

        file_size = os.path.getsize(ticker_file)
        return (ticker, dates, num_dates, available_metrics, file_size)

    except Exception as e:
        return (ticker, None, 0, [], 0, str(e))


def open_workbook(excel_path):
    """Open workbook with appropriate library based on extension.
    Returns (adapter, sheet_names, use_openpyxl)."""
    ext = os.path.splitext(excel_path)[1].lower()
    use_openpyxl = ext in ('.xlsx', '.xls', '.xlsm')

    if use_openpyxl:
        wb_ox = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        class SheetAdapter:
            def __init__(self, ws):
                self._ws = ws
            def __enter__(self):
                return self
            def __exit__(self, *a):
                pass
            def rows(self):
                for row in self._ws.iter_rows(values_only=False):
                    class Cell:
                        def __init__(self, v):
                            self.v = v
                    yield [Cell(c.value) for c in row]
        class WbAdapter:
            def __init__(self, ox):
                self._ox = ox
                self.sheets = ox.sheetnames
            def get_sheet(self, name):
                return SheetAdapter(self._ox[name])
            def close(self):
                self._ox.close()
        return WbAdapter(wb_ox), wb_ox.sheetnames, True
    else:
        reader = FastXlsbReader(excel_path)
        return reader, reader.sheet_names, False


def parse_workbook_to_dir(excel_path, output_dir):
    """Parse a workbook into a temp directory structure (same format as data/)."""
    t_start = time.time()
    os.makedirs(os.path.join(output_dir, "tickers"), exist_ok=True)
    
    wb, sheets, use_openpyxl = open_workbook(excel_path)

    mktdata_sheets = [s for s in sheets if '_mktdata' in s.lower()]
    print(f"Found {len(mktdata_sheets)} ticker sheets", file=sys.stderr)

    # Extract Tickerlist
    ticker_meta = {}
    if 'Tickerlist' in sheets:
        if use_openpyxl:
            with wb.get_sheet('Tickerlist') as sheet:
                for i, row in enumerate(sheet.rows()):
                    if i == 0:
                        continue
                    vals = [c.v for c in row[:8]]
                    if not vals[0]:
                        continue
                    raw_ticker = str(vals[0]).replace("-US", "").strip()
                    ticker_meta[raw_ticker] = {
                        "ticker": raw_ticker,
                        "name": str(vals[1] or "").strip(),
                        "economy": str(vals[2] or "").strip(),
                        "sector": str(vals[3] or "").strip(),
                        "subsector": str(vals[4] or "").strip(),
                        "industryGroup": str(vals[5] or "").strip(),
                        "industry": str(vals[6] or "").strip(),
                        "subindustry": str(vals[7] or "Other").strip(),
                    }
        else:
            tl_rows = wb.read_sheet_rows('Tickerlist', needed_rows=None, max_row=200)
            for r0 in sorted(tl_rows.keys()):
                if r0 == 0:
                    continue
                vals = tl_rows[r0]
                if len(vals) < 8 or not vals[0]:
                    continue
                raw_ticker = str(vals[0]).replace("-US", "").strip()
                ticker_meta[raw_ticker] = {
                    "ticker": raw_ticker,
                    "name": str(vals[1] or "").strip() if len(vals) > 1 else "",
                    "economy": str(vals[2] or "").strip() if len(vals) > 2 else "",
                    "sector": str(vals[3] or "").strip() if len(vals) > 3 else "",
                    "subsector": str(vals[4] or "").strip() if len(vals) > 4 else "",
                    "industryGroup": str(vals[5] or "").strip() if len(vals) > 5 else "",
                    "industry": str(vals[6] or "").strip() if len(vals) > 6 else "",
                    "subindustry": str(vals[7] or "Other").strip() if len(vals) > 7 else "Other",
                }

    # Extract Events
    events = {}
    for sheet_name, event_key in [("Ex dividend dates", "ex_dividend"), ("Earnings report dates", "earnings")]:
        if sheet_name not in sheets:
            continue
        if use_openpyxl:
            with wb.get_sheet(sheet_name) as sheet:
                for i, row in enumerate(sheet.rows()):
                    if i == 0:
                        continue
                    vals = [c.v for c in row]
                    if not vals[0]:
                        continue
                    ticker = str(vals[0]).replace("-US^", "").replace("-US", "").strip()
                    dates_list = []
                    for v in vals[1:]:
                        d = parse_date(v)
                        if d:
                            dates_list.append(d)
                    if ticker not in events:
                        events[ticker] = {}
                    events[ticker][event_key] = dates_list
        else:
            ev_rows = wb.read_sheet_rows(sheet_name, needed_rows=None, max_row=200)
            for r0 in sorted(ev_rows.keys()):
                if r0 == 0:
                    continue
                vals = ev_rows[r0]
                if not vals or not vals[0]:
                    continue
                ticker = str(vals[0]).replace("-US^", "").replace("-US", "").strip()
                dates_list = []
                for v in vals[1:]:
                    d = parse_date(v)
                    if d:
                        dates_list.append(d)
                if ticker not in events:
                    events[ticker] = {}
                events[ticker][event_key] = dates_list

    # Close workbook before parallel processing
    wb.close()

    # Extract Ticker Data (PARALLEL)
    tickers_dir = os.path.join(output_dir, "tickers")
    all_dates = None
    tickers_list = []

    import multiprocessing
    num_workers = min(multiprocessing.cpu_count(), 4)
    print(f"  Using {num_workers} parallel workers for {len(mktdata_sheets)} sheets", file=sys.stderr)

    work_args = [(excel_path, sn, tickers_dir) for sn in mktdata_sheets]

    with ProcessPoolExecutor(max_workers=num_workers) as pool:
        futures = {pool.submit(process_single_sheet, args): args[1] for args in work_args}
        completed = 0
        for future in as_completed(futures):
            completed += 1
            result = future.result()
            if len(result) > 5:
                ticker, _, _, _, _, err = result
                print(f"  [{completed}/{len(mktdata_sheets)}] {ticker} ERROR: {err}", file=sys.stderr)
                continue

            ticker, dates, num_dates, available_metrics, file_size = result
            if dates is None:
                print(f"  [{completed}/{len(mktdata_sheets)}] {ticker} SKIP", file=sys.stderr)
                continue

            print(f"  [{completed}/{len(mktdata_sheets)}] {ticker}: {len(available_metrics)} metrics, {file_size//1024}KB", file=sys.stderr)

            if all_dates is None or len(dates) > len(all_dates):
                all_dates = dates

            meta = ticker_meta.get(ticker, {
                "ticker": ticker, "name": ticker, "economy": "", "sector": "",
                "subsector": "", "industryGroup": "", "industry": "Other", "subindustry": "Other",
            })
            meta["dates"] = num_dates
            meta["metrics"] = available_metrics
            tickers_list.append(meta)

    # Write temp output files
    if all_dates:
        with open(os.path.join(output_dir, "dates.json"), 'w') as f:
            json.dump(all_dates, f, separators=(',', ':'))

    with open(os.path.join(output_dir, "tickers.json"), 'w') as f:
        json.dump(tickers_list, f, separators=(',', ':'))

    with open(os.path.join(output_dir, "events.json"), 'w') as f:
        json.dump(events, f, separators=(',', ':'))

    elapsed = time.time() - t_start
    print(f"  Parsing complete in {elapsed:.1f}s", file=sys.stderr)

    return tickers_list, events


def cmd_preview(excel_path, temp_dir):
    """Phase 1: Parse new workbook to temp dir and detect conflicts with existing data."""
    new_tickers, new_events = parse_workbook_to_dir(excel_path, temp_dir)
    new_ticker_names = set(t["ticker"] for t in new_tickers)

    # Load existing tickers
    existing_tickers_file = os.path.join(DATA_DIR, "tickers.json")
    existing_ticker_names = set()
    if os.path.exists(existing_tickers_file):
        try:
            existing = json.load(open(existing_tickers_file))
            existing_ticker_names = set(t["ticker"] for t in existing)
        except Exception:
            pass

    conflicts = sorted(new_ticker_names & existing_ticker_names)
    new_only = sorted(new_ticker_names - existing_ticker_names)

    result = {
        "newTickers": new_only,
        "conflicts": conflicts,
        "totalNew": len(new_tickers),
        "totalExisting": len(existing_ticker_names),
        "newEventTickers": len(new_events),
    }
    # Output to stdout as JSON
    print(json.dumps(result))


def cmd_merge(temp_dir, resolutions_json):
    """Phase 2: Merge temp data into main data/ dir with conflict resolutions."""
    resolutions = json.loads(resolutions_json)  # {"SPG": "overwrite", "O": "keep"}

    os.makedirs(os.path.join(DATA_DIR, "tickers"), exist_ok=True)

    # ── Load existing data ──
    existing_tickers = []
    existing_tickers_file = os.path.join(DATA_DIR, "tickers.json")
    if os.path.exists(existing_tickers_file):
        try:
            existing_tickers = json.load(open(existing_tickers_file))
        except Exception:
            pass
    existing_ticker_map = {t["ticker"]: t for t in existing_tickers}

    existing_events = {}
    existing_events_file = os.path.join(DATA_DIR, "events.json")
    if os.path.exists(existing_events_file):
        try:
            existing_events = json.load(open(existing_events_file))
        except Exception:
            pass

    existing_dates = []
    existing_dates_file = os.path.join(DATA_DIR, "dates.json")
    if os.path.exists(existing_dates_file):
        try:
            existing_dates = json.load(open(existing_dates_file))
        except Exception:
            pass

    # ── Load new (temp) data ──
    new_tickers = json.load(open(os.path.join(temp_dir, "tickers.json")))
    new_events = json.load(open(os.path.join(temp_dir, "events.json")))
    new_dates_file = os.path.join(temp_dir, "dates.json")
    new_dates = json.load(open(new_dates_file)) if os.path.exists(new_dates_file) else []

    # ── Load workbook sources tracking ──
    sources_file = os.path.join(DATA_DIR, "sources.json")
    sources = {}
    if os.path.exists(sources_file):
        try:
            sources = json.load(open(sources_file))
        except Exception:
            pass

    # Determine the workbook name for tracking
    wb_name = f"workbook-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    upload_ts = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

    # ── Merge dates (use the longer list) ──
    merged_dates = existing_dates if len(existing_dates) >= len(new_dates) else new_dates

    # ── Merge tickers ──
    merged_count = 0
    overwritten_count = 0
    skipped_count = 0

    for new_meta in new_tickers:
        ticker = new_meta["ticker"]
        new_ticker_file = os.path.join(temp_dir, "tickers", f"{ticker}.json")

        if ticker in existing_ticker_map:
            # Conflict — check resolution
            resolution = resolutions.get(ticker, "keep")
            if resolution == "overwrite":
                # Overwrite: copy new file and update metadata
                if os.path.exists(new_ticker_file):
                    import shutil
                    dest = os.path.join(DATA_DIR, "tickers", f"{ticker}.json")
                    shutil.copy2(new_ticker_file, dest)
                    fsize = os.path.getsize(dest)
                else:
                    fsize = 0
                existing_ticker_map[ticker] = new_meta
                sources[ticker] = {
                    "workbook": wb_name,
                    "uploadedAt": upload_ts,
                    "dates": new_meta.get("dates", 0),
                    "metrics": len(new_meta.get("metrics", [])),
                    "fileSize": fsize,
                }
                overwritten_count += 1
                print(f"  {ticker}: OVERWRITTEN", file=sys.stderr)
            else:
                # Keep existing
                skipped_count += 1
                print(f"  {ticker}: KEPT existing", file=sys.stderr)
        else:
            # New ticker — always add
            if os.path.exists(new_ticker_file):
                import shutil
                dest = os.path.join(DATA_DIR, "tickers", f"{ticker}.json")
                shutil.copy2(new_ticker_file, dest)
                fsize = os.path.getsize(dest)
            else:
                fsize = 0
            existing_ticker_map[ticker] = new_meta
            sources[ticker] = {
                "workbook": wb_name,
                "uploadedAt": upload_ts,
                "dates": new_meta.get("dates", 0),
                "metrics": len(new_meta.get("metrics", [])),
                "fileSize": fsize,
            }
            merged_count += 1
            print(f"  {ticker}: ADDED (new)", file=sys.stderr)

    # ── Merge events ──
    for ticker, evt in new_events.items():
        if ticker in existing_ticker_map:
            resolution = resolutions.get(ticker, "keep") if ticker in {t["ticker"] for t in new_tickers if t["ticker"] in {t2["ticker"] for t2 in existing_tickers}} else "add"
            if ticker not in existing_events or resolution == "overwrite" or resolution == "add":
                existing_events[ticker] = evt

    # ── Write merged output ──
    merged_tickers_list = sorted(existing_ticker_map.values(), key=lambda t: t["ticker"])

    with open(os.path.join(DATA_DIR, "dates.json"), 'w') as f:
        json.dump(merged_dates, f, separators=(',', ':'))

    with open(os.path.join(DATA_DIR, "tickers.json"), 'w') as f:
        json.dump(merged_tickers_list, f, separators=(',', ':'))

    with open(os.path.join(DATA_DIR, "events.json"), 'w') as f:
        json.dump(existing_events, f, separators=(',', ':'))

    with open(sources_file, 'w') as f:
        json.dump(sources, f, separators=(',', ':'), indent=2)

    result = {
        "totalTickers": len(merged_tickers_list),
        "added": merged_count,
        "overwritten": overwritten_count,
        "kept": skipped_count,
        "dates": len(merged_dates),
    }
    print(json.dumps(result))


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: merge-workbook.py preview|merge <args...>", file=sys.stderr)
        sys.exit(1)

    command = sys.argv[1]
    if command == "preview":
        if len(sys.argv) < 4:
            print("Usage: merge-workbook.py preview <excel_path> <temp_dir>", file=sys.stderr)
            sys.exit(1)
        cmd_preview(sys.argv[2], sys.argv[3])
    elif command == "merge":
        if len(sys.argv) < 4:
            print("Usage: merge-workbook.py merge <temp_dir> <resolutions_json>", file=sys.stderr)
            sys.exit(1)
        cmd_merge(sys.argv[2], sys.argv[3])
    else:
        print(f"Unknown command: {command}", file=sys.stderr)
        sys.exit(1)
